from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import workbook
import pathlib

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook



backgroud = "#234E70"
framebg = "#EDEDED"
framefg = "#101820"



root = Tk()
root.title("Student registration System")
root.geometry("1250x700+210+100")
root.configure(bg=backgroud)

file = pathlib.Path("Student_data.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration no."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Fathers Name"
    sheet['J1'] = "Mothers Name"
    sheet['K1'] = "Fathers Occupation"
    sheet['L1'] = "Mothers Occupation"

    file.save('Student_data.xlsx')


######################### Exit###############################
def Exit():
    root.destroy()

#########################Showimage###############################
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file", filetypes=(
    ("JPG File", "*.jpg"), ("PNG File", "*.png"), ("ALL Files", "*.txt")))

    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.configure(image=photo2)
    lbl.image = photo2
############################Registration no##################################(Automatic)
def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    Sheet = file.active
    row = Sheet.max_row

    max_row_value = Sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")

##############################Clear#####################################################
def Clear():
    Name.set('')
    Category.set('')
    Skills.set('')
    F_name.set('')
    M_name.set('')
    F_occupation.set('')
    M_occupation.set('')
    Class.set("Select Class")

    registration_no()

    saveButton.configure(state='normal')

    img1=PhotoImage(file='Images/upload photo.png')
    lbl.configure(image=img1)
    lbl.image=img1

    img=""

#######################Save########################################

def Save():
    global gender
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender")
    D2=DOB.get()
    D1=Date.get()
    CA1=Category.get()
    S1 = Skills.get()
    fathername=F_name.get()
    mothername=M_name.get()
    F1=F_occupation.get()
    M1=M_occupation.get()


    if  N1==""   or C1=="Select Class" or D2=="" or S1=="" or CA1=="" or fathername==""or mothername==""or F1=="" or M1=="":
        messagebox.showerror("error","few data is missing")
    else:
        file=openpyxl.load_workbook('Student_data.xlsx')
        Sheet=file.active
        Sheet.cell(column=1,row=Sheet.max_row+1,value=R1)
        Sheet.cell(column=2,row=Sheet.max_row,value=N1)
        Sheet.cell(column=3, row=Sheet.max_row, value=C1)
        Sheet.cell(column=4, row=Sheet.max_row, value=G1)
        Sheet.cell(column=5, row=Sheet.max_row, value=D2)
        Sheet.cell(column=6, row=Sheet.max_row, value=D1)
        Sheet.cell(column=7, row=Sheet.max_row, value=CA1)
        Sheet.cell(column=8, row=Sheet.max_row, value=S1)
        Sheet.cell(column=9, row=Sheet.max_row, value=fathername)
        Sheet.cell(column=10, row=Sheet.max_row, value=mothername)
        Sheet.cell(column=11, row=Sheet.max_row, value=F1)
        Sheet.cell(column=12, row=Sheet.max_row, value=M1)

        file.save(r'Student_data.xlsx')

        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not avalible !!! ")

        messagebox.showinfo("info","Sucessfully data entered !!")

        Clear()

        registration_no()

############################Searech#######################################

def search():
    global x1
    text=Search.get()
    Clear()
    saveButton.configure(state='disable')

    file=openpyxl.load_workbook("Student_data.xlsx")
    Sheet=file.active

    for row in Sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_no=str(name)[15:-1]
    try:
        print(str(name))
    except:
        messagebox.showerror("invalid","invalid registration number !!!")

    x1 = Sheet.cell(row=int(reg_no),column=1).value
    x2 = Sheet.cell(row=int(reg_no), column=2).value
    x3 = Sheet.cell(row=int(reg_no), column=3).value
    x4 = Sheet.cell(row=int(reg_no), column=4).value
    x5 = Sheet.cell(row=int(reg_no), column=5).value
    x6 = Sheet.cell(row=int(reg_no), column=6).value
    x7 = Sheet.cell(row=int(reg_no), column=7).value
    x8 = Sheet.cell(row=int(reg_no), column=8).value
    x9 = Sheet.cell(row=int(reg_no), column=9).value
    x10 = Sheet.cell(row=int(reg_no), column=10).value
    x11 = Sheet.cell(row=int(reg_no), column=11).value
    x12 = Sheet.cell(row=int(reg_no), column=12).value

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)
    if x4=="Female":
        R2.select()
    else:
        R1.select()
    DOB.set(x5)
    Date.set(x6)
    Category.set(x7)
    Skills.set(x8)
    F_name.set(x9)
    M_name.set(x10)
    F_occupation.set(x11)
    M_occupation.set(x12)

    img=(Image.open("Student Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.configure(image=photo2)
    lbl.img=photo2


################################Update####################################

def Update():
    global R1
    global gender
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    CA1 = Category.get()
    S1 = Skills.get()
    fathername = F_name.get()
    mothername = M_name.get()
    F1 = F_occupation.get()
    M1 = M_occupation.get()

    file = openpyxl.load_workbook('Student_data.xlsx')
    Sheet = file.active

    for row in Sheet.rows:
        if row[0].value == R1:
            name = row[0]
            print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_no = str(name)[15:-1]
    Sheet.cell(column=1, row=int(reg_no), value=R1)
    Sheet.cell(column=2, row=int(reg_no), value=N1)
    Sheet.cell(column=3, row=int(reg_no), value=C1)
    Sheet.cell(column=4, row=int(reg_no), value=G1)
    Sheet.cell(column=5, row=int(reg_no), value=D2)
    Sheet.cell(column=6, row=int(reg_no), value=D1)
    Sheet.cell(column=7, row=int(reg_no), value=CA1)
    Sheet.cell(column=8, row=int(reg_no), value=S1)
    Sheet.cell(column=9, row=int(reg_no), value=fathername)
    Sheet.cell(column=10, row=int(reg_no), value=mothername)
    Sheet.cell(column=11, row=int(reg_no), value=F1)
    Sheet.cell(column=12, row=int(reg_no), value=M1)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("Update","Update Sucessfully !!")
    Clear()



#################################Gender###################################
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "female"



# top frames

Label(root, text="Email:connectwithshashank@hotmail.com", width=10, height=3, bg="#234E70",fg="white", anchor='e').pack(side=TOP,
                                                                                                             fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="lightblue", fg='#234E70', font='arial 20 bold italic').pack(
    side=TOP, fill=X)

##########################Searchbox to update##################################################

Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font='arial 20').place(x=820, y=70)
imageicon3 = PhotoImage(file='Images/search.png')
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg="#234E70" ,fg='white', font="arial 13 bold",command=search)
Srch.place(x=1060, y=66)

imageicon4 = PhotoImage(file="Images/Layer 4.png")
bold_font = ("Calibri", 12, "bold")
update = Button(root,text="Update",font=bold_font,fg="white",bg="#234E70",width=15,command=Update)
update.place(x=110, y=72)

# #####################registration and date###############################
Label(root, text="Registration No:", font="arial 13", fg='#ffffff', bg=backgroud).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg='#ffffff', bg=backgroud).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()

# registration no()
today = date.today()
d1 = today.strftime(" %d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

Date.set(d1)

# student details
obj = LabelFrame(root, text="Student details", font="Calibri 17", bd=2, width=900, bg='#EDEDED', fg=framefg, height=250,
                 relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date of birth", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)

Label(obj, text="Course", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Category", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Skills", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

DOB = StringVar()
DOB_entry = Entry(obj, textvariable=DOB, width=20, font="arial 10")
DOB_entry.place(x=160, y=100)

Category = StringVar()
Category_entry = Entry(obj, textvariable=Category, width=20, font="arial 10")
Category_entry.place(x=630, y=100)

Skills = StringVar()
Skills_entry = Entry(obj, textvariable=Skills, width=20, font="arial 10")
Skills_entry.place(x=630, y=150)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

Class = Combobox(obj, values=["B.TECH", "MCA", "MBA", "M.TECH"], font="Roboto 10", width=17, state="r")
Class.place(x=630, y=50)
Class.set("Select Course")

########################### parents details##########################################################
obj2 = LabelFrame(root, text="Parent Details", font="Calibri 17", bd=2, width=900, bg=framebg, fg=framefg, height=250,
                  relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Father's Name", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Occupation", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

F_name = StringVar()
f_entry = Entry(obj2, textvariable=F_name, width=20, font="arial 10")
f_entry.place(x=160, y=50)

F_occupation = StringVar()
fo_entry = Entry(obj2, textvariable=F_occupation, width=20, font="arial 10")
fo_entry.place(x=160, y=100)

Label(obj2, text="Mother's Name", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Occupation", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

M_name = StringVar()
m_entry = Entry(obj2, textvariable=M_name, width=20, font="arial 10")
m_entry.place(x=630, y=50)

M_occupation = StringVar()
mo_entry = Entry(obj2, textvariable=M_occupation, width=20, font="arial 10")
mo_entry.place(x=630, y=100)

# image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="Images/upload photo.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# Button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000,
                                                                                                               y=370)

saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen",command=Save)
saveButton.place(x=1000, y=450)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink",command=Clear).place(x=1000, y=530)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="gray", command=Exit).place(x=1000, y=610)

root.mainloop()
